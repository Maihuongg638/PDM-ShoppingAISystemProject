# importing openpyxl module
from os import name
import openpyxl
import pymysql
import sys
	
host = "localhost"
user = "root"
passwd = ""
database = "shop"
def InsertItemToCart(UserName, id):
    connection = pymysql.connect(host=host , user=user, passwd=passwd, database=database)
    cursor = connection.cursor()
    retrive = "	INSERT INTO `product`(`ID`, `NAME`, `IMAGE`, `PRICE`, `RATING`, `CATEGORY_ID`, `DESCRIPTON`) VALUES ('[value-1]','[value-2]','[value-3]','[value-4]','[value-5]','[value-6]','[value-7]')".format(name = UserName, id = id)

    print(retrive)
	#executing the quires
    cursor.execute(retrive)
    connection.commit()
    return True

def removecharater(string):
	if '-' in string:
		string = string[0:string.find('-')]
		if '$' in string:
			string.replace('$', '')
		return string[1:len(string)]
# Give the location of the file
path = "data2.xlsx"
  
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
  
sheet_obj = wb_obj.active
  
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
  
# Will print a particular row value
for i in range(2, max_row + 1):
	print("-----------------------")
	cell_obj = sheet_obj.cell(row = i, column = 2)
	name = cell_obj.value
	name.replace("\'","\\'")
	print("Name = {name}".format(name = name))
	cell_obj = sheet_obj.cell(row = i, column = 3)
	print("Image = {name}".format(name = cell_obj.value))
	img = cell_obj.value
	cell_obj = sheet_obj.cell(row = i, column = 4)
	res = str(cell_obj.value)
	if '-' in res:
		res = removecharater(res)
	if res[0] == '$':
		res = res[1:len(res)]
	print("Price = {name}".format(name = res))
	price = res
	cell_obj = sheet_obj.cell(row = i, column = 5)
	print("Rating = {name}".format(name = cell_obj.value))
	rating = cell_obj.value
	cell_obj = sheet_obj.cell(row = i, column = 6)
	print("Category = {name}".format(name = cell_obj.value))
	cate = cell_obj.value
	cell_obj = sheet_obj.cell(row = i, column = 7)
	des = str(cell_obj.value)
	print("Des = {name}".format(name = des))
	print("-----------------------")

	connection = pymysql.connect(host=host , user=user, passwd=passwd, database=database)
	cursor = connection.cursor()
	retrive = "INSERT INTO `product`(`ID`, `NAME`, `IMAGE`, `PRICE`, `RATING`, `CATEGORY_ID`, `DESCRIPTON`) VALUES ('','{name}','{img}','{price}','{rating}','{cate}','{des}')".format(name = name, img = img, price = price, rating = rating, cate = cate, des = des)
	print(retrive)
	#executing the quires
	cursor.execute(retrive)
	connection.commit()
	